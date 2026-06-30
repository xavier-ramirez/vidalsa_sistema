import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Datos combinados calculados:
datos_combinados = [
    ["VOLTEO", 2, 9, 11],
    ["CAMIONETA", 4, 8, 12],
    ["PAYLOADER", 0, 7, 7],
    ["CUADRILLERA", 0, 1, 1],
    ["CAMION DE SOLDADURA", 0, 1, 1],
    ["CHUTO", 16, 2, 18],
    ["EXCAVADORA", 0, 7, 7],
    ["LOWBOY", 2, 1, 3],
    ["MOTONIVELADORA", 0, 2, 2],
    ["RETROEXCAVADORA", 0, 2, 2],
    ["LUMINARIA", 0, 4, 4],
    ["VACUUM", 10, 2, 12],
    ["CAMION CON BRAZO 4 TON", 1, 0, 1],
    ["CAMION ELEVADOR 8 TON", 1, 0, 1],
    ["AMBULANCIA", 1, 1, 2],
    ["TRACTOR DE ORUGA", 1, 3, 4],
    ["CHUTO CON BRAZO 16 TON", 1, 0, 1],
    ["CAMION DE SERVICIO", 1, 0, 1],
    ["TRACTOR AGRICOLA", 2, 0, 2]
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resumen Equipos"

# Estilos
header_fill = PatternFill("solid", fgColor="004C8C")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
data_font = Font(name="Arial", size=10)
c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
l_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="000000")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

# Encabezados
headers = ["TIPO DE EQUIPO", "NUEVO\n(≥ 2025)", "VIEJO\n(< 2025)", "TOTAL"]
ws.append(headers)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = c_align
    cell.border = brd
ws.row_dimensions[1].height = 30

# Agregar datos
for i, row in enumerate(datos_combinados):
    ws.append(row)
    current_row = i + 2
    for col_idx, cell in enumerate(ws[current_row], 1):
        cell.font = data_font
        cell.border = brd
        cell.alignment = l_align if col_idx == 1 else c_align

# Ancho de columnas
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

# Guardar
out_path = r'C:\Users\dell12\Desktop\vidalsa_sistema\RESUMEN_EQUIPOS_COMBINADO.xlsx'
wb.save(out_path)
print(f"Archivo guardado en: {out_path}")
