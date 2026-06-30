import re
from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Leer el DOCX
doc = Document(r'C:\Users\dell12\Videos\LISTADO DE MAQUINARIAS Y EQUIPOS ACTUALIZADA_020320.docx')

all_rows = []
for table in doc.tables:
    for row in table.rows:
        cells = [c.text.strip() for c in row.cells]
        if cells[0].upper() == 'ITEM':
            continue
        if not any(cells):
            continue
        all_rows.append(cells)

def split_equipment_row(row):
    item, tipo, marca, modelo, serial = row[0], row[1], row[2], row[3], row[4]
    separators = [r'\n', r'\r', r' / ', r'/', r' Y ']
    for sep in separators:
        partes_tipo = re.split(sep, tipo, flags=re.IGNORECASE)
        partes_serial = re.split(sep, serial, flags=re.IGNORECASE)
        if len(partes_tipo) > 1:
            result = []
            for idx, p_tipo in enumerate(partes_tipo):
                p_tipo = p_tipo.strip()
                if not p_tipo:
                    continue
                p_serial = partes_serial[idx].strip() if idx < len(partes_serial) else ''
                sub_item = f"{item}{'abcdefgh'[idx]}" if idx > 0 else item
                result.append([sub_item, p_tipo, marca, modelo, p_serial])
            return result
    return [row]

final_rows = []
for row in all_rows:
    final_rows.extend(split_equipment_row(row))

# Crear Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Listado de Equipos"

header_fill  = PatternFill("solid", fgColor="1F3864")
alt_fill     = PatternFill("solid", fgColor="DCE6F1")
white_fill   = PatternFill("solid", fgColor="FFFFFF")
header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
data_font    = Font(name="Calibri", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin   = Side(style="thin", color="B0B8C1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Titulo
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = "LISTADO DE MAQUINARIAS Y EQUIPOS"
title_cell.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
title_cell.fill  = PatternFill("solid", fgColor="1A237E")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Encabezados
headers = ["N.", "TIPO DE EQUIPO", "MARCA", "MODELO", "SERIAL / PLACA"]
ws.append(headers)
for col_num, cell in enumerate(ws[2], 1):
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = border
ws.row_dimensions[2].height = 22

# Datos
for i, row_data in enumerate(final_rows):
    ws.append(row_data)
    row_num = i + 3
    fill = alt_fill if i % 2 == 0 else white_fill
    for col_num, cell in enumerate(ws[row_num], 1):
        cell.font      = data_font
        cell.fill      = fill
        cell.border    = border
        cell.alignment = center_align if col_num == 1 else left_align
    ws.row_dimensions[row_num].height = 18

# Anchos
col_widths = {1: 7, 2: 28, 3: 18, 4: 16, 5: 25}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

output_path = r'C:\Users\dell12\Desktop\LISTADO_EQUIPOS_VIDALSA.xlsx'
wb.save(output_path)
print(f"Excel generado con {len(final_rows)} equipos.")
print(f"Guardado en: {output_path}")
