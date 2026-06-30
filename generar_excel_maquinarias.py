import win32com.client as win32
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Leer doc via COM ─────────────────────────────────────────────────────────
src = 'C:\\Users\\dell12\\Downloads\\LISTADO DE MAQUINARIAS EQUIPOS (1).doc'
word = win32.Dispatch('Word.Application')
word.Visible = False
word.DisplayAlerts = False
doc = word.Documents.Open(src)

raw_rows = []
largest_tbl = None
max_rows = 0
for i in range(1, doc.Tables.Count + 1):
    if doc.Tables(i).Rows.Count > max_rows:
        max_rows = doc.Tables(i).Rows.Count
        largest_tbl = doc.Tables(i)

if largest_tbl:
    for r in range(2, largest_tbl.Rows.Count + 1):   # saltar encabezado (fila 1)
        row_data = []
        for c in range(1, largest_tbl.Columns.Count + 1):
            try:
                txt = largest_tbl.Cell(r, c).Range.Text.strip()
                txt = txt.rstrip('\r').rstrip('\x07')
            except:
                txt = ''
            row_data.append(txt)
        raw_rows.append(row_data)

doc.Close(False)
word.Quit()

# Columnas: [FLOTA, DESCRIPCION, PLACA, SERIAL_CHASIS, SERIAL_MOTOR, RESPONSABLE]

# ── Separar filas que tienen dos equipos (chuto + lowboy) ────────────────────
final_rows = []
item = 1

for r in raw_rows:
    # Pad the row to have at least 6 elements in case of merged cells or bad rows
    while len(r) < 6:
        r.append('')
    flota, desc, placa, serial_ch, serial_mot, resp = r[:6]

    # Si tiene serial de CHASIS y serial de MOTOR (son dos equipos distintos: chuto + lowboy)
    if serial_ch and serial_mot and ('CHUTO' in desc.upper() or 'LOWBOY' in desc.upper()):
        # Fila 1: el CHUTO (usa la placa y el serial de chasis)
        final_rows.append([item, flota, desc, placa, serial_ch, '', resp])
        item += 1
        # Fila 2: el LOW BOY (sin placa propia registrada, usa serial de motor como serial)
        final_rows.append([item, flota, 'LOW BOY', '', serial_mot, '', resp])
        item += 1
    else:
        # Filas normales: LOW BOY independiente ya listados — no duplicar
        # Saltar filas LOW BOY que ya fueron separadas (seriales que ya aparecieron)
        final_rows.append([item, flota, desc, placa, serial_ch, serial_mot, resp])
        item += 1

# Filtrar LOW BOY duplicados (los que ya se separaron arriba)
seriales_lowboy_separados = {'LJRL163C4F2004853', 'LJRL163CXF2004842', 'LJRL163C1F2004874'}

deduped_rows = []
for row in final_rows:
    desc_row = row[2]
    serial_row = row[4]
    # Si es un LOW BOY independiente con serial ya separado, saltarlo
    if desc_row == 'LOW BOY' and serial_row in seriales_lowboy_separados:
        continue
    deduped_rows.append(row)

# Recalcular numeración limpia
for i, row in enumerate(deduped_rows):
    row[0] = i + 1

# ── Crear Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Listado de Maquinarias"

# Estilos
h_fill  = PatternFill("solid", fgColor="1F3864")
alt1    = PatternFill("solid", fgColor="DCE6F1")
alt2    = PatternFill("solid", fgColor="FFFFFF")
sep_fill= PatternFill("solid", fgColor="E8F5E9")  # verde claro para LOW BOY separados

h_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
d_font  = Font(name="Calibri", size=9)
c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
l_align = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin    = Side(style="thin", color="AAAAAA")
brd     = Border(left=thin, right=thin, top=thin, bottom=thin)

# Título
ws.merge_cells("A1:G1")
tc = ws["A1"]
tc.value = "LISTADO DE MAQUINARIAS Y EQUIPOS"
tc.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
tc.fill  = PatternFill("solid", fgColor="0D1B2A")
tc.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Encabezados
headers = ["N°", "FLOTA", "DESCRIPCIÓN", "PLACA", "SERIAL CHASIS", "SERIAL MOTOR", "RESPONSABLE"]
ws.append(headers)
for cell in ws[2]:
    cell.font      = h_font
    cell.fill      = h_fill
    cell.alignment = c_align
    cell.border    = brd
ws.row_dimensions[2].height = 20

# Datos
for i, row in enumerate(deduped_rows):
    ws.append(row)
    rn   = i + 3
    desc = row[2]
    # LOW BOY separado = fondo verde claro para diferenciarlo visualmente
    if desc == 'LOW BOY' and row[4] in seriales_lowboy_separados:
        fill = sep_fill
    else:
        fill = alt1 if i % 2 == 0 else alt2

    for col_idx, cell in enumerate(ws[rn], 1):
        cell.font      = d_font
        cell.fill      = fill
        cell.border    = brd
        cell.alignment = c_align if col_idx in (1,) else l_align
    ws.row_dimensions[rn].height = 16

# Anchos
col_widths = {1:5, 2:14, 3:32, 4:16, 5:22, 6:22, 7:22}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Guardar
out = r'C:\Users\dell12\Desktop\vidalsa_sistema\LISTADO_MAQUINARIAS_EQUIPOS.xlsx'
wb.save(out)
print(f"Excel generado: {len(deduped_rows)} equipos")
print(f"Guardado en: {out}")
