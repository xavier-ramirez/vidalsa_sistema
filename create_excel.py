import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Columnas:
# TIPO | EQUIPO/CARGO | PLACA CHUTO | SERIAL CHUTO | PLACA LOWBOY/BATEA | SERIAL LOWBOY/BATEA | NOMBRE Y APELLIDO | C.I. | TELEFONO

data = [
    ["TIPO DE FLOTA", "EQUIPO / CARGO", "PLACA\nCHUTO", "SERIAL\nCHUTO", "PLACA\nLOWBOY / BATEA", "SERIAL\nLOWBOY / BATEA", "NOMBRE Y APELLIDO", "C.I.", "TELÉFONO"],
    # FLOTA PESADA - Volquetas (solo tienen placa y serial propio, sin lowboy)
    ["FLOTA PESADA", "VOLQUETA SINOTRUK",           "A29BM7F",    "LZZ1ELSF4SJ401914",    "",         "",                      "ALEXANDER MORALES",  "10.943.356", ""],
    ["FLOTA PESADA", "VOLQUETA SINOTRUK",           "A88BO4F",    "LZZ1ELSF9SJ413122",    "",         "",                      "CARMELO ECHENIQUE",  "18.893.892", ""],
    ["FLOTA PESADA", "VOLQUETA SINOTRUK (CARACAS)", "A06EA9G",    "LJ13R8DK4H3400100",    "",         "",                      "YVAN ROJAS",         "9.268.956",  ""],
    # CHUTOS CON LOWBOY
    ["FLOTA PESADA", "CHUTO CON LOWBOY SINOTRUK",  "A94BE5R",    "LZZPCMSC2SJ389421",    "A68ES8P",  "LJRL163C4F2004853",     "MARCO MENDEZ",       "9.691.753",  ""],
    ["FLOTA PESADA", "CHUTO CON LOWBOY SINOTRUK",  "A86BE6R",    "LZZPCMSC8SJ389424",    "",         "LJRL163CXF2004842",     "WILLIAM PALMERO",    "12.564.535", ""],
    ["FLOTA PESADA", "CHUTO CON LOWBOY SINOTRUK",  "A94BE8R",    "LZZPCMSC3SJ389217",    "A43ES3P",  "LJRL163C1F2004874",     "WILSON FLORES",      "12.876.777", ""],
    ["FLOTA PESADA", "CHUTO CON LOWBOY",           "A86BE8R",    "",                     "A69ES3P",  "LJRL13372F2004769",     "VICTOR ALVAREZ",     "15.729.843", ""],
    ["FLOTA PESADA", "CHUTO CON LOWBOY",           "A46BN0R",    "",                     "S/P",      "LMN9FRT34S0001206",     "LUIS GOLINDANO",     "13.589.516", ""],
    ["FLOTA PESADA", "CHUTO CON LOWBOY SINOTRUK",  "A00BO5F",    "",                     "S/P",      "LMN9FRT33S0001214",     "FRANCISCO SARACUAL", "12.806.018", ""],
    # Maquinaria (sin lowboy)
    ["FLOTA PESADA", "RETROEXCAVADORA LOVOL",       "",           "CLW009LDHSZ000227",    "",         "",                      "ALI BLANCO",         "17.590.732", ""],
    ["FLOTA PESADA", "RETROEXCAVADORA LOVOL",       "",           "CLW009LDTSZ000233",    "",         "",                      "VICTOR ESTUPIÑÁN",   "20.195.059", ""],
    ["FLOTA PESADA", "RETROEXCAVADORA CAT 420XE",  "",           "CAT00420JH9X00159",    "",         "",                      "JOSÉ HERNANDEZ",     "8.296.298",  ""],
    ["FLOTA PESADA", "JUMBO 220 LOVOL",             "",           "FTC003RHASS556889",    "",         "",                      "JORGE CAÑAS",        "14.187.366", ""],
    ["FLOTA PESADA", "JUMBO LOVOL (MARTILLO/420)",  "",           "FTC003RNESZ555561",    "",         "",                      "JOSÉ GUZMÁN",        "8.491.373",  ""],
    ["FLOTA PESADA", "PAYLOADER LOVOL FL956H",      "",           "CLW009LFPSW001712",    "",         "",                      "ALFREDO MOYA",       "12.151.310", ""],
    ["FLOTA PESADA", "PAYLOADER LOVOL FL976H",      "",           "CLW009LHKSW002054",    "",         "",                      "GORGE MAITA",        "17.091.193", ""],
    ["FLOTA PESADA", "TRACTOR D6 SHANTUI SD22",     "",           "CHSD22AAAS1028407",    "",         "",                      "JUNIOR BRITO",       "18.206.755", ""],
    ["FLOTA PESADA", "CHUTO BRAZO 16 TON SINOTRUK","A35BK1R",    "",                     "",         "",                      "CARLOS PEREZ",       "20.548.159", ""],
    ["FLOTA PESADA", "CAMION GRÚA 20 TON",          "A85BE6R",    "",                     "",         "",                      "OVIDIO CLEMANT",     "11.343.312", ""],
    ["FLOTA PESADA", "CAMIÓN DE SERVICIO",           "A51EX9P",    "",                     "",         "",                      "ISAIAS LARA",        "20.002.939", "0412-823.49.00"],
    # FLOTA LIVIANA
    ["FLOTA LIVIANA", "CAMIONETA SINOTRUK",          "A60X0I",     "",                     "",         "",                      "LEOSEL POITO",       "17.902.185", "0414-854.91.02"],
    ["FLOTA LIVIANA", "CAMIONETA SINOTRUK",          "A85BO0R",    "",                     "",         "",                      "HÉCTOR GARCÍA",      "15.803.090", "0424-891.79.27"],
    ["FLOTA LIVIANA", "CAMIONETA SINOTRUK",          "A82BN5F",    "",                     "",         "",                      "JOSÉ DOMINGUEZ",     "12.834.129", "0414-136.03.86"],
    ["FLOTA LIVIANA", "CAMIONETA SINOTRUK",          "A82B01R",    "",                     "",         "",                      "LUIS RONDON",        "21.390.029", "0412-833.03.53"],
    # PERSONAL
    ["PERSONAL", "COORDINADOR GENERAL",             "N/A",        "N/A",                  "",         "",                      "JOSE COVA",          "9.281.917",  "0414-768.69.44"],
    ["PERSONAL", "MECÁNICO",                        "N/A",        "N/A",                  "",         "",                      "GROBER PORTILLO",    "16.799.899", "0424-840.48.76"],
    ["PERSONAL", "MECÁNICO",                        "N/A",        "N/A",                  "",         "",                      "HENRY AZÓCAR",       "25.452.214", "0414-184.67.93"],
    ["PERSONAL", "AYUDANTE MECÁNICO",               "N/A",        "N/A",                  "",         "",                      "JESÚS GONZÁLEZ",     "31.295.817", "0412-729.5518"],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Recursos Emergencia"

# --- Estilos ---
header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill   = PatternFill("solid", fgColor="1F497D")
pesada_fill   = PatternFill("solid", fgColor="DCE6F1")
liviana_fill  = PatternFill("solid", fgColor="EBF1DE")
personal_fill = PatternFill("solid", fgColor="FFF2CC")
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin   = Side(style="thin", color="A0A0A0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Encabezado ---
ws.append(data[0])
for cell in ws[1]:
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = border

# --- Datos ---
fill_map = {
    "FLOTA PESADA":  pesada_fill,
    "FLOTA LIVIANA": liviana_fill,
    "PERSONAL":      personal_fill,
}

for row_data in data[1:]:
    ws.append(row_data)
    row  = ws.max_row
    fill = fill_map.get(row_data[0], PatternFill())
    for col in range(1, 10):
        cell           = ws.cell(row=row, column=col)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = center_align if col in (1, 8, 9) else left_align

# --- Anchos de columna ---
col_widths = [16, 28, 14, 22, 14, 22, 26, 13, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Alturas ---
ws.row_dimensions[1].height = 35
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 22

# --- Congelar encabezado ---
ws.freeze_panes = "A2"

output_path = r"C:\Users\dell12\Videos\ASIGNACION_RECURSOS_EMERGENCIA_Tabla.xlsx"
wb.save(output_path)
print(f"Archivo guardado: {output_path}")
